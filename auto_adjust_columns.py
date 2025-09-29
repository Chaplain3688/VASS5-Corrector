from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def auto_adjust_columns(file):
    print('Creating file:', file)
    wb = load_workbook(file)
    ws = wb.active

    table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="RobotData", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",  # Or use Light1, Medium2, Dark3, etc.
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    for column_cells in ws.columns:
        # Get the max length of the values in each column (as strings)
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
        adjusted_width = max_length + 5  # padding
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width
